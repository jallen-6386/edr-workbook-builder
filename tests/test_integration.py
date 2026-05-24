"""
Integration tests for build_workbook: verifies the full assembly pipeline
produces a valid .xlsx with the correct sheets, tab colours, and data.
"""

from pathlib import Path

import pandas as pd
import pytest
from openpyxl import load_workbook

from edr_workbook_builder.csv_loader import CSVLoadResult
from edr_workbook_builder.workbook import build_workbook


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _make_result(data: dict, name: str = "test.csv", error: str = None) -> CSVLoadResult:
    if error:
        return CSVLoadResult(
            path=Path(name), dataframe=None, encoding_used=None,
            error=error, row_count=0, col_count=0,
        )
    df = pd.DataFrame(data)
    return CSVLoadResult(
        path=Path(name), dataframe=df, encoding_used="utf-8",
        error=None, row_count=len(df), col_count=len(df.columns),
    )


def _sheet_headers(wb, sheet_name: str) -> list[str]:
    ws = wb[sheet_name]
    return [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]


# ---------------------------------------------------------------------------
# Basic workbook creation
# ---------------------------------------------------------------------------


class TestBasicWorkbook:
    def test_creates_xlsx_file(self, tmp_path):
        r = _make_result({"Col": ["a", "b"]})
        out = tmp_path / "out.xlsx"
        build_workbook([r], ["s1"], [None], out)
        assert out.exists()

    def test_correct_sheet_names(self, tmp_path):
        r1 = _make_result({"Col": ["a"]}, "a.csv")
        r2 = _make_result({"Col": ["b"]}, "b.csv")
        out = tmp_path / "out.xlsx"
        build_workbook([r1, r2], ["alpha", "beta"], [None, None], out)
        wb = load_workbook(out)
        assert "alpha" in wb.sheetnames
        assert "beta" in wb.sheetnames

    def test_errored_results_skipped(self, tmp_path):
        r_bad  = _make_result({}, error="parse error")
        r_good = _make_result({"Col": ["a"]})
        out = tmp_path / "out.xlsx"
        build_workbook([r_bad, r_good], ["bad", "good"], [None, None], out)
        wb = load_workbook(out)
        assert "good" in wb.sheetnames
        assert "bad" not in wb.sheetnames

    def test_no_data_creates_fallback_sheet(self, tmp_path):
        out = tmp_path / "out.xlsx"
        build_workbook([], [], [], out)
        wb = load_workbook(out)
        assert "No Data" in wb.sheetnames

    def test_header_row_has_all_columns(self, tmp_path):
        r = _make_result({"A": ["1"], "B": ["2"], "C": ["3"]})
        out = tmp_path / "out.xlsx"
        build_workbook([r], ["data"], [None], out)
        wb = load_workbook(out)
        assert _sheet_headers(wb, "data") == ["A", "B", "C"]


# ---------------------------------------------------------------------------
# Summary sheet
# ---------------------------------------------------------------------------


class TestSummarySheet:
    def test_summary_sheet_created(self, tmp_path):
        r = _make_result({"Col": ["a"]})
        out = tmp_path / "out.xlsx"
        build_workbook([r], ["data"], [None], out, add_summary=True)
        wb = load_workbook(out)
        assert "Analysis_Summary" in wb.sheetnames

    def test_summary_is_first_sheet(self, tmp_path):
        r = _make_result({"Col": ["a"]})
        out = tmp_path / "out.xlsx"
        build_workbook([r], ["data"], [None], out, add_summary=True)
        wb = load_workbook(out)
        assert wb.sheetnames[0] == "Analysis_Summary"


# ---------------------------------------------------------------------------
# Timeline sheet
# ---------------------------------------------------------------------------


class TestTimelineSheet:
    def test_timeline_sheet_created(self, tmp_path):
        r = _make_result({"Timestamp": ["2024-01-01T00:00:00Z"], "Event": ["A"]})
        out = tmp_path / "out.xlsx"
        build_workbook([r], ["data"], [None], out, add_timeline=True)
        wb = load_workbook(out)
        assert "Timeline" in wb.sheetnames

    def test_timeline_has_green_tab(self, tmp_path):
        r = _make_result({"Timestamp": ["2024-01-01T00:00:00Z"], "Event": ["A"]})
        out = tmp_path / "out.xlsx"
        build_workbook([r], ["data"], [None], out, add_timeline=True)
        wb = load_workbook(out)
        assert "70AD47" in wb["Timeline"].sheet_properties.tabColor.rgb.upper()


# ---------------------------------------------------------------------------
# Process tree sheet
# ---------------------------------------------------------------------------


class TestProcessTreeSheet:
    def test_process_tree_created(self, tmp_path):
        r = _make_result({
            "ProcessId": ["100", "200"],
            "ParentProcessId": ["0", "100"],
            "ImageFileName": ["a.exe", "b.exe"],
        })
        out = tmp_path / "out.xlsx"
        build_workbook([r], ["data"], [None], out, add_process_tree=True)
        wb = load_workbook(out)
        assert "ProcessTree" in wb.sheetnames

    def test_process_tree_has_purple_tab(self, tmp_path):
        r = _make_result({
            "ProcessId": ["100"],
            "ParentProcessId": ["0"],
            "ImageFileName": ["a.exe"],
        })
        out = tmp_path / "out.xlsx"
        build_workbook([r], ["data"], [None], out, add_process_tree=True)
        wb = load_workbook(out)
        assert "7030A0" in wb["ProcessTree"].sheet_properties.tabColor.rgb.upper()

    def test_process_tree_node_count(self, tmp_path):
        r = _make_result({
            "ProcessId": ["100", "200"],
            "ParentProcessId": ["0", "100"],
        })
        out = tmp_path / "out.xlsx"
        result = build_workbook([r], ["data"], [None], out, add_process_tree=True)
        assert result.process_tree_node_count == 2


# ---------------------------------------------------------------------------
# ATT&CK column
# ---------------------------------------------------------------------------


class TestAttckColumn:
    def test_attck_column_in_headers(self, tmp_path):
        r = _make_result({
            "ImageFileName": ["powershell.exe"],
            "CommandLine":   ["whoami"],
        })
        out = tmp_path / "out.xlsx"
        build_workbook([r], ["data"], [None], out, add_attck=True)
        wb = load_workbook(out)
        assert "ATT&CK" in _sheet_headers(wb, "data")

    def test_attck_column_after_process_col(self, tmp_path):
        r = _make_result({
            "ImageFileName": ["powershell.exe"],
            "CommandLine":   ["whoami"],
            "Extra":         ["x"],
        })
        out = tmp_path / "out.xlsx"
        build_workbook([r], ["data"], [None], out, add_attck=True)
        wb = load_workbook(out)
        headers = _sheet_headers(wb, "data")
        assert headers.index("ATT&CK") == headers.index("ImageFileName") + 1

    def test_attck_cell_contains_technique_name(self, tmp_path):
        r = _make_result({
            "ImageFileName": ["powershell.exe"],
            "CommandLine":   [""],
        })
        out = tmp_path / "out.xlsx"
        build_workbook([r], ["data"], [None], out, add_attck=True)
        wb = load_workbook(out)
        ws = wb["data"]
        headers = _sheet_headers(wb, "data")
        attck_col = headers.index("ATT&CK") + 1
        cell_val = ws.cell(2, attck_col).value
        assert "PowerShell" in cell_val


# ---------------------------------------------------------------------------
# Decoded command column
# ---------------------------------------------------------------------------


class TestDecodedCommand:
    def _encode(self, cmd: str) -> str:
        import base64
        return base64.b64encode(cmd.encode("utf-16-le")).decode("ascii")

    def test_decoded_column_appears(self, tmp_path):
        blob = self._encode("whoami")
        r = _make_result({
            "ImageFileName": ["powershell.exe"],
            "CommandLine":   [f"-EncodedCommand {blob}"],
        })
        out = tmp_path / "out.xlsx"
        build_workbook([r], ["data"], [None], out, decode_encoded=True)
        wb = load_workbook(out)
        assert "DecodedCommand" in _sheet_headers(wb, "data")

    def test_decoded_column_after_commandline(self, tmp_path):
        blob = self._encode("Get-Process")
        r = _make_result({
            "ImageFileName": ["powershell.exe"],
            "CommandLine":   [f"-enc {blob}"],
            "Extra":         ["x"],
        })
        out = tmp_path / "out.xlsx"
        build_workbook([r], ["data"], [None], out, decode_encoded=True)
        wb = load_workbook(out)
        headers = _sheet_headers(wb, "data")
        assert headers.index("DecodedCommand") == headers.index("CommandLine") + 1

    def test_no_encoded_command_no_column(self, tmp_path):
        r = _make_result({
            "ImageFileName": ["cmd.exe"],
            "CommandLine":   ["whoami"],
        })
        out = tmp_path / "out.xlsx"
        build_workbook([r], ["data"], [None], out, decode_encoded=True)
        wb = load_workbook(out)
        assert "DecodedCommand" not in _sheet_headers(wb, "data")


# ---------------------------------------------------------------------------
# IOC extract sheet
# ---------------------------------------------------------------------------


class TestIocExtractSheet:
    def test_ioc_sheet_created(self, tmp_path):
        r = _make_result({"SHA256": ["a" * 64], "RemoteIP": ["1.2.3.4"]})
        out = tmp_path / "out.xlsx"
        build_workbook([r], ["data"], [None], out, add_ioc_sheet=True)
        wb = load_workbook(out)
        assert "IOC_Extract" in wb.sheetnames

    def test_ioc_sheet_has_orange_tab(self, tmp_path):
        r = _make_result({"SHA256": ["a" * 64]})
        out = tmp_path / "out.xlsx"
        build_workbook([r], ["data"], [None], out, add_ioc_sheet=True)
        wb = load_workbook(out)
        assert "ED7D31" in wb["IOC_Extract"].sheet_properties.tabColor.rgb.upper()

    def test_ioc_count_in_result(self, tmp_path):
        r = _make_result({"SHA256": ["a" * 64, "b" * 64], "RemoteIP": ["1.2.3.4", ""]})
        out = tmp_path / "out.xlsx"
        result = build_workbook([r], ["data"], [None], out, add_ioc_sheet=True)
        assert result.ioc_count == 3

    def test_no_ioc_columns_no_sheet(self, tmp_path):
        r = _make_result({"Col": ["value"]})
        out = tmp_path / "out.xlsx"
        build_workbook([r], ["data"], [None], out, add_ioc_sheet=True)
        wb = load_workbook(out)
        assert "IOC_Extract" not in wb.sheetnames


# ---------------------------------------------------------------------------
# columns_filter
# ---------------------------------------------------------------------------


class TestColumnsFilter:
    def test_only_specified_columns_written(self, tmp_path):
        r = _make_result({"A": ["1"], "B": ["2"], "C": ["3"]})
        out = tmp_path / "out.xlsx"
        build_workbook([r], ["data"], [None], out, columns_filter=["A", "C"])
        wb = load_workbook(out)
        headers = _sheet_headers(wb, "data")
        assert "A" in headers
        assert "C" in headers
        assert "B" not in headers

    def test_nonexistent_column_filter_graceful(self, tmp_path):
        r = _make_result({"A": ["1"]})
        out = tmp_path / "out.xlsx"
        # Filters that match nothing fall through without error.
        build_workbook([r], ["data"], [None], out, columns_filter=["NonExistent"])
        wb = load_workbook(out)
        # When no columns match the filter, the full df is kept.
        assert "A" in _sheet_headers(wb, "data")


# ---------------------------------------------------------------------------
# max_rows
# ---------------------------------------------------------------------------


class TestMaxRows:
    def test_rows_truncated(self, tmp_path):
        r = _make_result({"Col": [str(i) for i in range(50)]})
        out = tmp_path / "out.xlsx"
        build_workbook([r], ["data"], [None], out, max_rows=10)
        wb = load_workbook(out)
        ws = wb["data"]
        # Row 1 = header, rows 2–11 = 10 data rows.
        assert ws.max_row == 11

    def test_rows_not_truncated_when_under_limit(self, tmp_path):
        r = _make_result({"Col": [str(i) for i in range(5)]})
        out = tmp_path / "out.xlsx"
        build_workbook([r], ["data"], [None], out, max_rows=100)
        wb = load_workbook(out)
        ws = wb["data"]
        assert ws.max_row == 6  # header + 5 data rows


# ---------------------------------------------------------------------------
# Sheet order with multiple special sheets
# ---------------------------------------------------------------------------


class TestSheetOrder:
    def test_summary_first_when_all_enabled(self, tmp_path):
        r = _make_result({
            "Timestamp":       ["2024-01-01T00:00:00Z"],
            "ProcessId":       ["100"],
            "ParentProcessId": ["0"],
            "ImageFileName":   ["cmd.exe"],
            "SHA256":          ["a" * 64],
        })
        out = tmp_path / "out.xlsx"
        build_workbook(
            [r], ["data"], [None], out,
            add_summary=True, add_timeline=True,
            add_process_tree=True, add_ioc_sheet=True,
        )
        wb = load_workbook(out)
        assert wb.sheetnames[0] == "Analysis_Summary"
