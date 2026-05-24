"""
Analysis_Summary sheet builder.

Creates the first worksheet with:
  - Case metadata and processing statistics
  - Worksheet inventory (with LOLBin flag column, added v0.3)
  - Column inventory matrix (added v0.2)
  - Parent/child process relationship table (added v0.2)
  - Suspicious Activity section (added v0.3, only when --highlight is used)
  - Analyst notes area
  - Suggested Excel Copilot prompts
"""

from datetime import datetime
from typing import Optional

from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from edr_workbook_builder.patterns import LOLBINS, RowFinding, extract_exe_name
from edr_workbook_builder.process_detection import extract_exe_name

_TITLE_FONT = Font(name="Calibri", bold=True, size=14, color="1F497D")
_SECTION_FONT = Font(name="Calibri", bold=True, size=11, color="1F497D")
_LABEL_FONT = Font(name="Calibri", bold=True, size=10)
_VALUE_FONT = Font(name="Calibri", size=10)
_TABLE_HDR_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
_TABLE_HDR_FILL = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
_WARN_FONT = Font(name="Calibri", size=10, color="C00000")
_CHECK_FONT = Font(name="Calibri", size=10, color="375623")
_ALT_FILL = PatternFill(start_color="EEF2F7", end_color="EEF2F7", fill_type="solid")

# Severity label fonts for the Suspicious Activity table.
_SEVERITY_FONTS = {
    1: Font(name="Calibri", size=10, color="7D6608"),   # dark yellow — LOLBin
    2: Font(name="Calibri", size=10, color="7E5109"),   # dark orange — obfuscation
    3: Font(name="Calibri", size=10, color="922B21"),   # dark red    — encoded cmd
}
_SEVERITY_LABELS = {1: "Info (LOLBin)", 2: "Medium", 3: "High"}

_COPILOT_PROMPTS = [
    "Summarize suspicious process activity across all sheets in this workbook.",
    "Identify unusual or suspicious command-line arguments.",
    "Find all network connections, file writes, registry modifications, or encoded commands.",
    "Compare parent and child process relationships and identify anomalies.",
    "Identify indicators that suggest benign activity versus suspicious activity.",
    "List all unique processes and their frequency of occurrence.",
    (
        "Identify LOLBin usage: powershell, rundll32, regsvr32, mshta, certutil, "
        "wscript, cscript, bitsadmin, schtasks, cmd."
    ),
    "Look for base64-encoded or otherwise obfuscated command-line arguments.",
    "Build a chronological timeline of events using available timestamp fields.",
    "Identify lateral movement indicators such as remote host connections or admin share access.",
]

# Key EDR field groups checked in the column inventory matrix.
_KEY_EDR_GROUPS: dict[str, list[str]] = {
    "Timestamp": ["Timestamp", "EventTimeUTC", "ContextTimeStamp", "EventTime", "StartTime"],
    "ProcessId": ["ProcessId", "Pid", "ProcessId64"],
    "ImageFileName": ["ImageFileName", "FileName", "ProcessName"],
    "CommandLine": ["CommandLine"],
    "ParentProcessId": ["ParentProcessId", "ParentPid", "ParentProcessId64"],
    "ParentProcess": ["ParentBaseFileName", "ParentProcessImageFileName", "ParentImageFileName"],
    "Network": ["RemoteAddressIP4", "RemoteAddressIP6", "RemotePort", "RemoteIP", "LocalPort"],
    "FileTarget": ["TargetFileName", "TargetFilePath"],
    "Hash": ["SHA256HashData", "MD5HashData", "SHA1HashData"],
    "Registry": ["TargetObjectName"],
}

_PARENT_PID_COLS = ["ParentProcessId", "ParentPid", "ParentProcessId64"]
_CHILD_PID_COLS = ["ProcessId", "Pid", "ProcessId64"]
_PARENT_NAME_COLS = ["ParentBaseFileName", "ParentProcessImageFileName", "ParentImageFileName"]
_CHILD_NAME_COLS = ["ImageFileName", "FileName", "ProcessName"]

_MAX_RELATIONSHIPS = 200
_MAX_SUSPICIOUS_ROWS = 500


def _find_col(columns_lower: dict[str, str], candidates: list[str]) -> Optional[str]:
    for c in candidates:
        if c.lower() in columns_lower:
            return columns_lower[c.lower()]
    return None


def _section(ws, row: int, title: str) -> int:
    ws.cell(row=row, column=1, value=title).font = _SECTION_FONT
    return row + 1


def _kv(ws, row: int, label: str, value: str = "", warn: bool = False) -> int:
    ws.cell(row=row, column=1, value=label).font = _LABEL_FONT
    cell = ws.cell(row=row, column=2, value=value)
    cell.font = _WARN_FONT if warn else _VALUE_FONT
    return row + 1


def _table_header(ws, row: int, headers: list[str]) -> int:
    for c_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=c_idx, value=h)
        cell.font = _TABLE_HDR_FONT
        cell.fill = _TABLE_HDR_FILL
        cell.alignment = Alignment(horizontal="left")
    return row + 1


def get_key_column_presence(columns: list[str]) -> dict[str, bool]:
    """Return which key EDR groups are present in a list of column names."""
    columns_lower = {c.lower() for c in columns}
    return {
        group: any(cand.lower() in columns_lower for cand in candidates)
        for group, candidates in _KEY_EDR_GROUPS.items()
    }


def extract_relationships(load_results: list, sheet_names: list[str]) -> list[dict]:
    """
    Extract unique parent/child process relationships from loaded CSVs.

    Returns a list of dicts with keys: parent_name, parent_pid, child_name,
    child_pid, source_sheet. Only populated when PID columns are present.
    Capped at _MAX_RELATIONSHIPS rows to keep the summary sheet manageable.
    """
    seen: set[tuple[str, str]] = set()
    relationships: list[dict] = []

    for result, sheet_name in zip(load_results, sheet_names):
        if result.dataframe is None or result.error is not None:
            continue

        df = result.dataframe
        cols_lower = {c.lower(): c for c in df.columns}

        parent_pid_col = _find_col(cols_lower, _PARENT_PID_COLS)
        child_pid_col = _find_col(cols_lower, _CHILD_PID_COLS)
        if not (parent_pid_col and child_pid_col):
            continue

        parent_name_col = _find_col(cols_lower, _PARENT_NAME_COLS)
        child_name_col = _find_col(cols_lower, _CHILD_NAME_COLS)

        cols_to_use = [c for c in [parent_pid_col, child_pid_col, parent_name_col, child_name_col] if c]
        for _, row_data in df[cols_to_use].drop_duplicates().iterrows():
            if len(relationships) >= _MAX_RELATIONSHIPS:
                break

            parent_pid = str(row_data.get(parent_pid_col, "")).strip()
            child_pid = str(row_data.get(child_pid_col, "")).strip()

            if not parent_pid or not child_pid or parent_pid == "nan" or child_pid == "nan":
                continue

            key = (parent_pid, child_pid)
            if key in seen:
                continue
            seen.add(key)

            raw_parent = str(row_data.get(parent_name_col, "")).strip() if parent_name_col else ""
            raw_child = str(row_data.get(child_name_col, "")).strip() if child_name_col else ""

            parent_name = (extract_exe_name(raw_parent) or raw_parent) if raw_parent and raw_parent != "nan" else ""
            child_name = (extract_exe_name(raw_child) or raw_child) if raw_child and raw_child != "nan" else ""

            relationships.append({
                "parent_name": parent_name or "unknown",
                "parent_pid": parent_pid,
                "child_name": child_name or "unknown",
                "child_pid": child_pid,
                "source_sheet": sheet_name,
            })

    return relationships


def _write_column_inventory(ws, row: int, load_results: list, sheet_names: list[str]) -> int:
    """Write per-sheet column inventory with key EDR field presence matrix."""
    row = _section(ws, row, "COLUMN INVENTORY")

    key_groups = list(_KEY_EDR_GROUPS.keys())
    headers = ["Sheet", "Rows", "Cols"] + key_groups + ["All Columns"]
    row = _table_header(ws, row, headers)

    for i, (result, sheet_name) in enumerate(zip(load_results, sheet_names)):
        if result.dataframe is None or result.error is not None:
            continue

        df = result.dataframe
        presence = get_key_column_presence(list(df.columns))
        fill = _ALT_FILL if (i % 2 == 1) else None

        def _cell(col_idx, value, font=_VALUE_FONT, align=None):
            c = ws.cell(row=row, column=col_idx, value=value)
            c.font = font
            if fill:
                c.fill = fill
            if align:
                c.alignment = align

        _cell(1, sheet_name)
        _cell(2, result.row_count)
        _cell(3, len(df.columns))

        for j, group_name in enumerate(key_groups, start=4):
            found = presence[group_name]
            c = ws.cell(row=row, column=j, value="✓" if found else "")
            c.font = _CHECK_FONT if found else _VALUE_FONT
            c.alignment = Alignment(horizontal="center")
            if fill:
                c.fill = fill

        all_cols = ", ".join(df.columns.tolist())
        if len(all_cols) > 250:
            all_cols = all_cols[:247] + "..."
        _cell(4 + len(key_groups), all_cols)

        row += 1

    return row + 1


def _write_relationship_table(ws, row: int, load_results: list, sheet_names: list[str]) -> int:
    """Write parent/child process relationship table (only when PID columns are present)."""
    relationships = extract_relationships(load_results, sheet_names)
    if not relationships:
        return row

    row = _section(ws, row, "PARENT / CHILD PROCESS RELATIONSHIPS")
    headers = ["Parent Process", "Parent PID", "Child Process", "Child PID", "Source Sheet"]
    row = _table_header(ws, row, headers)

    for i, rel in enumerate(relationships):
        fill = _ALT_FILL if (i % 2 == 1) else None
        for c_idx, val in enumerate(
            [rel["parent_name"], rel["parent_pid"], rel["child_name"], rel["child_pid"], rel["source_sheet"]],
            start=1,
        ):
            c = ws.cell(row=row, column=c_idx, value=val)
            c.font = _VALUE_FONT
            if fill:
                c.fill = fill
        row += 1

    if len(relationships) == _MAX_RELATIONSHIPS:
        ws.cell(row=row, column=1, value=f"(truncated at {_MAX_RELATIONSHIPS} rows)").font = _WARN_FONT
        row += 1

    return row + 1


def _write_suspicious_activity(ws, row: int, findings: list[RowFinding]) -> int:
    """Write Suspicious Activity summary (only when --highlight produced findings)."""
    if not findings:
        return row

    row = _section(ws, row, "SUSPICIOUS ACTIVITY SUMMARY")

    # Stats
    high = sum(1 for f in findings if f.severity >= 3)
    medium = sum(1 for f in findings if f.severity == 2)
    lolbin_only = sum(1 for f in findings if f.severity == 1)

    row = _kv(ws, row, "Total flagged rows:", str(len(findings)), warn=bool(high or medium))
    if high:
        row = _kv(ws, row, "High (encoded command):", str(high), warn=True)
    if medium:
        row = _kv(ws, row, "Medium (obfuscation):", str(medium), warn=True)
    if lolbin_only:
        row = _kv(ws, row, "Info (LOLBin process):", str(lolbin_only))

    # Color legend
    ws.cell(row=row, column=1, value="Color key:").font = _LABEL_FONT
    ws.cell(row=row, column=2,
            value="Yellow = LOLBin process  |  Salmon = possible obfuscation  |  Red = encoded command"
            ).font = _VALUE_FONT
    row += 2

    # Findings table
    capped = findings[:_MAX_SUSPICIOUS_ROWS]
    headers = ["Sheet", "Row #", "Process", "Pattern Detected", "Severity"]
    row = _table_header(ws, row, headers)

    for i, f in enumerate(capped):
        fill = _ALT_FILL if (i % 2 == 1) else None
        sev_label = _SEVERITY_LABELS.get(f.severity, str(f.severity))
        sev_font = _SEVERITY_FONTS.get(f.severity, _VALUE_FONT)

        for c_idx, val in enumerate(
            [f.sheet_name, f.data_row, f.process_name or "(unknown)", "; ".join(f.reasons)],
            start=1,
        ):
            c = ws.cell(row=row, column=c_idx, value=val)
            c.font = _VALUE_FONT
            if fill:
                c.fill = fill

        c = ws.cell(row=row, column=5, value=sev_label)
        c.font = sev_font
        if fill:
            c.fill = fill

        row += 1

    if len(findings) > _MAX_SUSPICIOUS_ROWS:
        ws.cell(row=row, column=1,
                value=f"(showing first {_MAX_SUSPICIOUS_ROWS} of {len(findings)} flagged rows)"
                ).font = _WARN_FONT
        row += 1

    return row + 1


def create_summary_sheet(
    ws,
    case_name: Optional[str],
    timestamp: datetime,
    load_results: list,
    sheet_names: list[str],
    process_names: list[Optional[str]],
    suspicious_findings: Optional[list[RowFinding]] = None,
) -> None:
    successful = [r for r in load_results if r.error is None and r.dataframe is not None]
    failed = [r for r in load_results if r.error is not None]
    total_rows = sum(r.row_count for r in successful)

    row = 1
    ws.cell(row=row, column=1, value="EDR Analysis Workbook — Summary").font = _TITLE_FONT
    row += 2

    row = _section(ws, row, "CASE INFORMATION")
    row = _kv(ws, row, "Case / Alert:", case_name or "(not provided)")
    row = _kv(ws, row, "Generated:", timestamp.strftime("%Y-%m-%d %H:%M:%S"))
    row = _kv(ws, row, "Tool:", "EDR Workbook Builder v0.3.0")
    row += 1

    row = _section(ws, row, "PROCESSING STATISTICS")
    row = _kv(ws, row, "CSV files found:", str(len(load_results)))
    row = _kv(ws, row, "Files processed:", str(len(successful)))
    row = _kv(ws, row, "Files skipped (errors):", str(len(failed)), warn=bool(failed))
    row = _kv(ws, row, "Total data rows:", f"{total_rows:,}")
    row = _kv(ws, row, "Worksheets created:", str(len(successful)))
    if suspicious_findings is not None:
        row = _kv(ws, row, "Suspicious rows flagged:", str(len(suspicious_findings)),
                  warn=bool(suspicious_findings))
    row += 1

    # Worksheet inventory (v0.3 adds LOLBin? column)
    row = _section(ws, row, "WORKSHEET INVENTORY")
    headers = ["Sheet Name", "Source CSV", "Detected Process", "Row Count", "Status", "LOLBin?"]
    row = _table_header(ws, row, headers)

    for result, sheet_name, proc_name in zip(load_results, sheet_names, process_names):
        status = "OK" if result.error is None else f"SKIPPED — {result.error}"
        is_err = result.error is not None
        is_lb = bool(proc_name and proc_name.lower() in LOLBINS)

        ws.cell(row=row, column=1, value=sheet_name).font = _VALUE_FONT
        ws.cell(row=row, column=2, value=result.path.name).font = _VALUE_FONT
        ws.cell(row=row, column=3, value=proc_name or "(fallback: filename)").font = _VALUE_FONT
        ws.cell(row=row, column=4, value=result.row_count).font = _VALUE_FONT
        ws.cell(row=row, column=5, value=status).font = _WARN_FONT if is_err else _VALUE_FONT
        lolbin_cell = ws.cell(row=row, column=6, value="⚠ Yes" if is_lb else "")
        lolbin_cell.font = _WARN_FONT if is_lb else _VALUE_FONT
        if is_lb:
            lolbin_cell.alignment = Alignment(horizontal="center")
        row += 1
    row += 1

    # v0.2: column inventory matrix
    row = _write_column_inventory(ws, row, load_results, sheet_names)

    # v0.2: parent/child relationship table
    row = _write_relationship_table(ws, row, load_results, sheet_names)

    # v0.3: suspicious activity summary (only when --highlight was used)
    if suspicious_findings is not None:
        row = _write_suspicious_activity(ws, row, suspicious_findings)

    # Import errors
    if failed:
        row = _section(ws, row, "IMPORT ERRORS / WARNINGS")
        for result in failed:
            row = _kv(ws, row, result.path.name, result.error or "Unknown error", warn=True)
        row += 1

    # Analyst notes
    row = _section(ws, row, "ANALYST NOTES")
    for i in range(1, 7):
        ws.cell(row=row, column=1, value=f"Note {i}:").font = _LABEL_FONT
        ws.cell(row=row, column=2, value="").font = _VALUE_FONT
        row += 1
    row += 1

    # Copilot prompts
    row = _section(ws, row, "SUGGESTED EXCEL COPILOT PROMPTS")
    for prompt in _COPILOT_PROMPTS:
        ws.cell(row=row, column=1, value=f"• {prompt}").font = _VALUE_FONT
        row += 1

    # Column widths — accommodate the widest section (column inventory at 14 cols).
    key_group_count = len(_KEY_EDR_GROUPS)
    col_widths: dict[int, int] = {
        1: 28,   # Sheet Name / Parent Process / labels
        2: 46,   # Source CSV / Parent PID / values
        3: 24,   # Detected Process / Child Process / Cols
        4: 14,   # Row Count / Child PID / first key group
        5: 50,   # Status / Source Sheet
        6: 12,   # LOLBin? (v0.3) / third key group
    }
    # Key group columns in the inventory matrix (columns 4 through 4+key_group_count-1).
    for i in range(key_group_count):
        col_num = 4 + i
        if col_num not in col_widths or col_widths[col_num] < 14:
            col_widths[col_num] = 14
    # "All Columns" is at column 4 + key_group_count.
    col_widths[4 + key_group_count] = 60

    for col_num, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col_num)].width = width
