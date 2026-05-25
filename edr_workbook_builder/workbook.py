"""
Excel workbook construction and per-sheet formatting.

Formatting choices:
  - Dark blue (#1F497D) header row with white bold text — easy to scan
  - Freeze pane at A2 so the header stays visible while scrolling
  - Auto-filter on every sheet so analysts can filter without extra steps
  - Column width sampled from up to 500 rows and capped at 60 chars
    (prevents CommandLine columns from making the sheet unusably wide)
  - All values written as strings — no type inference that could mangle
    hashes, timestamps, or numeric PIDs
  - Optional row highlighting for suspicious patterns (--highlight):
      yellow  = LOLBin process name (severity 1)
      salmon  = possible obfuscation, base64, or hex blob (severity 2)
      red     = PowerShell -EncodedCommand detected (severity 3)
  - Optional formula escaping (--escape-formulas): cell values starting
    with =, +, -, or @ are prefixed with ' to prevent formula injection
"""

import logging
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

from edr_workbook_builder.patterns import (
    FORMULA_PFXS,
    RowFinding,
    SuspiciousMatch,
    check_row,
    max_severity,
)
from edr_workbook_builder.timeline import (
    build_timeline_df,
    find_best_timestamp_column,
)


@dataclass
class WorkbookResult:
    findings: list[RowFinding] = field(default_factory=list)
    timeline_row_count: int = 0
    timeline_included: list[str] = field(default_factory=list)
    timeline_excluded: list[str] = field(default_factory=list)
    timeline_timestamp_col: Optional[str] = None
    process_tree_node_count: int = 0
    ioc_count: int = 0
    decoded_command_count: int = 0  # rows where DecodedCommand was populated

logger = logging.getLogger(__name__)

_HEADER_FILL = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
_HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
_HEADER_ALIGN = Alignment(horizontal="left", vertical="center")
_BODY_FONT = Font(name="Calibri", size=10)
_BODY_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=False)

# Row fill colors by severity level.
_SEVERITY_FILLS = {
    1: PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),  # yellow  - LOLBin
    2: PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"),  # salmon  - obfuscation
    3: PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid"),  # red     - encoded cmd
}

_MAX_COL_WIDTH = 60
_MIN_COL_WIDTH = 8
_SAMPLE_ROWS = 500


def _auto_size_columns(ws, df: pd.DataFrame) -> None:
    for col_idx, col_name in enumerate(df.columns, start=1):
        col_letter = get_column_letter(col_idx)
        max_len = len(str(col_name))
        for val in df.iloc[:_SAMPLE_ROWS, col_idx - 1]:
            if val:
                vlen = len(str(val))
                if vlen > max_len:
                    max_len = vlen
        ws.column_dimensions[col_letter].width = min(
            max(max_len + 2, _MIN_COL_WIDTH), _MAX_COL_WIDTH
        )


def write_dataframe_to_sheet(
    ws,
    df: pd.DataFrame,
    add_source: bool = False,
    source_name: str = "",
    highlight_suspicious: bool = False,
    escape_formulas: bool = False,
) -> list[tuple[int, list[SuspiciousMatch]]]:
    """
    Write a DataFrame to an openpyxl worksheet with standard EDR formatting.

    Returns a list of (1-indexed data row number, suspicious matches) for any
    rows that matched suspicious patterns. Empty list when highlight_suspicious
    is False or no matches were found.
    """
    if df.empty and len(df.columns) == 0:
        ws["A1"] = "(no data in source CSV)"
        ws["A1"].font = Font(name="Calibri", italic=True, color="808080")
        return []

    if add_source and source_name:
        df = df.copy()
        df.insert(0, "SourceFile", source_name)

    df = df.fillna("")

    # Pre-compute pattern matches for every data row when highlighting is on.
    # Done up front so the writing loop stays linear.
    row_matches_list: list[list[SuspiciousMatch]] = []
    if highlight_suspicious:
        for _, row in df.iterrows():
            row_matches_list.append(check_row(row))
    else:
        row_matches_list = [[] for _ in range(len(df))]

    findings: list[tuple[int, list[SuspiciousMatch]]] = []
    data_row_idx = 0

    for r_idx, row_vals in enumerate(dataframe_to_rows(df, index=False, header=True), start=1):
        if r_idx == 1:
            for c_idx, value in enumerate(row_vals, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                cell.font = _HEADER_FONT
                cell.fill = _HEADER_FILL
                cell.alignment = _HEADER_ALIGN
        else:
            matches = row_matches_list[data_row_idx] if data_row_idx < len(row_matches_list) else []
            severity = max_severity(matches)
            row_fill = _SEVERITY_FILLS.get(severity)

            for c_idx, value in enumerate(row_vals, start=1):
                if escape_formulas and isinstance(value, str) and value and value[0] in FORMULA_PFXS:
                    value = "'" + value
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                cell.font = _BODY_FONT
                cell.alignment = _BODY_ALIGN
                if row_fill:
                    cell.fill = row_fill

            if matches:
                findings.append((data_row_idx + 1, matches))

            data_row_idx += 1

    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 18

    if ws.dimensions:
        ws.auto_filter.ref = ws.dimensions

    _auto_size_columns(ws, df)

    return findings


def build_workbook(
    load_results: list,
    sheet_names: list[str],
    process_names: list[Optional[str]],
    output_path: Path,
    case_name: Optional[str] = None,
    add_summary: bool = False,
    add_source_column: bool = False,
    timestamp: Optional[datetime] = None,
    highlight_suspicious: bool = False,
    escape_formulas: bool = False,
    add_timeline: bool = False,
    add_attck: bool = False,
    add_process_tree: bool = False,
    decode_encoded: bool = False,
    add_ioc_sheet: bool = False,
    columns_filter: Optional[list[str]] = None,
    max_rows: Optional[int] = None,
) -> WorkbookResult:
    """Assemble and save the Excel workbook from loaded CSV results.

    Returns a WorkbookResult containing suspicious row findings and timeline
    metadata (counts, included/excluded sheets, timestamp column used).
    """
    if timestamp is None:
        timestamp = datetime.now()

    wb = Workbook()
    wb.remove(wb.active)

    all_findings: list[RowFinding] = []
    sheets_created = 0
    decoded_count = 0

    for result, sheet_name in zip(load_results, sheet_names):
        if result.error or result.dataframe is None:
            continue

        df = result.dataframe

        # Column filter — applied before any augmentation so feature columns
        # (ATT&CK, DecodedCommand) are added on top of the filtered set.
        if columns_filter:
            keep = [c for c in df.columns if c in columns_filter]
            df = df[keep] if keep else df

        # Row cap.
        if max_rows is not None and len(df) > max_rows:
            logger.warning(
                "Sheet '%s': truncated to %d rows (had %d)", sheet_name, max_rows, len(df)
            )
            df = df.iloc[:max_rows]

        if add_attck:
            from edr_workbook_builder.attck import add_attck_column
            df = add_attck_column(df)

        if decode_encoded:
            from edr_workbook_builder.patterns import add_decoded_column
            df = add_decoded_column(df)
            if "DecodedCommand" in df.columns:
                decoded_count += (df["DecodedCommand"] != "").sum()

        ws = wb.create_sheet(title=sheet_name)
        sheet_findings = write_dataframe_to_sheet(
            ws,
            df,
            add_source=add_source_column,
            source_name=result.path.name,
            highlight_suspicious=highlight_suspicious,
            escape_formulas=escape_formulas,
        )
        sheets_created += 1
        logger.debug("Wrote sheet '%s' (%d rows)", sheet_name, result.row_count)

        for data_row, matches in sheet_findings:
            process_name = next((m.process_exe for m in matches if m.process_exe), "")
            all_findings.append(RowFinding(
                sheet_name=sheet_name,
                data_row=data_row,
                process_name=process_name,
                reasons=[m.reason for m in matches],
                severity=max_severity(matches),
            ))

    # Guarantee at least one sheet (Excel requires it).
    if sheets_created == 0 and not add_summary and not add_timeline:
        ws = wb.create_sheet(title="No Data")
        ws["A1"] = "No CSV files could be processed. Run with --verbose for details."
        ws["A1"].font = Font(name="Calibri", bold=True, color="C00000")

    result = WorkbookResult(findings=all_findings, decoded_command_count=decoded_count)

    # IOC Extract sheet — orange tab, inserted at index 0.
    if add_ioc_sheet:
        from edr_workbook_builder.ioc_extract import extract_iocs
        ioc_df = extract_iocs(load_results, sheet_names)
        if not ioc_df.empty:
            ws_ioc = wb.create_sheet(title="IOC_Extract", index=0)
            ws_ioc.sheet_properties.tabColor = "ED7D31"  # orange tab
            write_dataframe_to_sheet(ws_ioc, ioc_df)
            result.ioc_count = len(ioc_df)
            logger.info("IOC_Extract: %d unique indicator(s) written", len(ioc_df))
        else:
            logger.warning("IOC_Extract: no hash or IP columns found — sheet not created")

    # ProcessTree sheet — inserted at index 0.
    # Timeline (if requested) is then inserted at index 0, pushing ProcessTree to 1.
    # Summary (if requested) is inserted at index 0 last, pushing all others down.
    if add_process_tree:
        from edr_workbook_builder.proctree import build_process_tree_rows
        tree_rows = build_process_tree_rows(load_results, sheet_names)
        if tree_rows:
            tree_df = pd.DataFrame(tree_rows, columns=["Process", "PID", "PPID", "CommandLine", "SourceSheet"])
            ws_tree = wb.create_sheet(title="ProcessTree", index=0)
            ws_tree.sheet_properties.tabColor = "7030A0"  # purple tab
            write_dataframe_to_sheet(ws_tree, tree_df)
            result.process_tree_node_count = len(tree_rows)
            logger.info("ProcessTree: %d node(s) written", len(tree_rows))
        else:
            logger.warning(
                "ProcessTree: no ProcessId/ParentProcessId columns found — sheet not created"
            )

    # Timeline sheet — inserted at index 0, pushed to index 1 when summary is added.
    if add_timeline:
        ts_col = find_best_timestamp_column(load_results, sheet_names)
        if ts_col is None:
            logger.warning(
                "Timeline: no timestamp column found across loaded CSVs — "
                "sheet not created (looked for: Timestamp, EventTimeUTC, "
                "ContextTimeStamp, EventTime, StartTime, EndTime, CreationTime)",
            )
        else:
            tl_df, tl_included, tl_excluded = build_timeline_df(
                load_results, sheet_names, ts_col
            )
            if tl_df is not None:
                ws_tl = wb.create_sheet(title="Timeline", index=0)
                ws_tl.sheet_properties.tabColor = "70AD47"  # green tab for easy navigation
                write_dataframe_to_sheet(ws_tl, tl_df)
                result.timeline_timestamp_col = ts_col
                result.timeline_row_count = len(tl_df)
                result.timeline_included = tl_included
                result.timeline_excluded = tl_excluded
                logger.info(
                    "Timeline: %d rows from %d sheet(s) sorted by '%s'%s",
                    len(tl_df), len(tl_included), ts_col,
                    f" ({len(tl_excluded)} excluded — no '{ts_col}' column)" if tl_excluded else "",
                )
            else:
                logger.warning("Timeline: no sheets contained '%s' — sheet not created", ts_col)

    if add_summary:
        from edr_workbook_builder.summary import create_summary_sheet
        ws_summary = wb.create_sheet(title="Analysis_Summary", index=0)
        create_summary_sheet(
            ws=ws_summary,
            case_name=case_name,
            timestamp=timestamp,
            load_results=load_results,
            sheet_names=sheet_names,
            process_names=process_names,
            suspicious_findings=all_findings if highlight_suspicious else None,
        )

    wb.properties.title = case_name or "EDR Analysis Workbook"
    wb.properties.creator = "EDR Workbook Builder"
    wb.properties.description = f"Generated {timestamp.isoformat()}"

    wb.save(output_path)
    logger.info("Workbook saved: %s", output_path)

    return result
